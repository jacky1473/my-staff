import os
import io
import secrets
import sqlite3
import pytz
import logging
import random
import string
import calendar
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pdfplumber
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import shutil
import json

# ---------------------------------------------------------------------------
# App Setup
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-in-production-please')

DB_PATH = '/data/attendance.db' if os.path.exists('/data') else 'attendance.db'
BACKUP_DIR = os.environ.get('BACKUP_DIR', './backups')

# Ensure backup directory exists
os.makedirs(BACKUP_DIR, exist_ok=True)

ALLOWED_DEPARTMENTS = ['IT', 'MIS', 'QA', 'TL', 'Manager', 'Management']
SHIFT_OPTIONS = [
    '09:00 AM - 06:00 PM',
    '10:00 AM - 07:00 PM',
    '12:00 PM - 09:00 PM',
    '02:00 PM - 11:00 PM',
    'Night Shift',
    'Flexible',
]
WEEKDAY_OPTIONS = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

# Brute-force protection
_login_attempts = {}
MAX_ATTEMPTS    = 3  # Reduced from 5 to 3 for local network
LOCKOUT_MINUTES = 5   # Reduced from 15 to 5

# PIN storage: {username: {'pin': code, 'expires': datetime, 'role': 'Admin'|'Staff'}}
_pin_storage = {}
PIN_EXPIRY_MINUTES = 3  # 3 minute expiry

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------
def get_db_connection():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode=WAL")   # Better concurrent access
    conn.execute("PRAGMA synchronous=NORMAL") # Faster writes, still safe
    return conn


def init_db():
    conn = get_db_connection()

    conn.execute('''CREATE TABLE IF NOT EXISTS company (
        id   INTEGER PRIMARY KEY,
        name TEXT    NOT NULL
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS users (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        username          TEXT UNIQUE NOT NULL,
        password          TEXT NOT NULL,
        department        TEXT NOT NULL,
        role              TEXT NOT NULL,
        shift             TEXT DEFAULT '09:00 AM - 06:00 PM',
        weekoff           TEXT DEFAULT 'Sunday',
        security_question TEXT DEFAULT 'What is your favorite color?',
        security_answer   TEXT DEFAULT 'blue'
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id   INTEGER NOT NULL,
        date      TEXT    NOT NULL,
        clock_in  TEXT,
        clock_out TEXT,
        status    TEXT DEFAULT 'Present',
        FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS announcements (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        title      TEXT    NOT NULL,
        body       TEXT    NOT NULL,
        priority   TEXT    DEFAULT 'normal',
        created_at TEXT    NOT NULL,
        created_by TEXT    NOT NULL,
        active     INTEGER DEFAULT 1
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS audit_logs (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id   INTEGER,
        action    TEXT NOT NULL,
        details   TEXT,
        ip_addr   TEXT,
        timestamp TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE SET NULL
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS activity_logs (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id      INTEGER NOT NULL UNIQUE,
        username     TEXT NOT NULL,
        status       TEXT DEFAULT 'offline',
        current_page TEXT DEFAULT '/',
        ip_addr      TEXT,
        last_seen    TEXT,
        last_active  TEXT,
        FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS leave_requests (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER NOT NULL,
        username    TEXT NOT NULL,
        leave_type  TEXT NOT NULL,
        from_date   TEXT NOT NULL,
        to_date     TEXT NOT NULL,
        days        INTEGER NOT NULL DEFAULT 1,
        reason      TEXT NOT NULL,
        status      TEXT DEFAULT 'Pending',
        admin_note  TEXT DEFAULT '',
        applied_at  TEXT NOT NULL,
        reviewed_at TEXT DEFAULT '',
        reviewed_by TEXT DEFAULT '',
        FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS notifications (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER NOT NULL,
        title      TEXT NOT NULL,
        message    TEXT NOT NULL,
        type       TEXT DEFAULT 'info',
        is_read    INTEGER DEFAULT 0,
        created_at TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS leave_balance (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER NOT NULL UNIQUE,
        casual      INTEGER DEFAULT 12,
        sick        INTEGER DEFAULT 12,
        earned      INTEGER DEFAULT 15,
        FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
    )''')

    _safe_alter(conn, "ALTER TABLE attendance ADD COLUMN status TEXT DEFAULT 'Present'")
    _safe_alter(conn, "ALTER TABLE users ADD COLUMN shift TEXT DEFAULT '09:00 AM - 06:00 PM'")
    _safe_alter(conn, "ALTER TABLE users ADD COLUMN weekoff TEXT DEFAULT 'Sunday'")
    _safe_alter(conn, "ALTER TABLE users ADD COLUMN security_question TEXT DEFAULT 'What is your favorite color?'")
    _safe_alter(conn, "ALTER TABLE users ADD COLUMN security_answer TEXT DEFAULT 'blue'")
    _safe_alter(conn, "ALTER TABLE users ADD COLUMN full_name TEXT DEFAULT ''")

    conn.commit()
    conn.close()


def _safe_alter(conn, sql):
    try:
        conn.execute(sql)
    except Exception:
        pass

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def ist_now():
    return datetime.now(pytz.timezone('Asia/Kolkata'))


def today_str():
    return ist_now().strftime('%Y-%m-%d')


def get_todays_roster():
    today = today_str()
    conn  = get_db_connection()
    roster = conn.execute('''
        SELECT u.username, u.full_name, u.department, u.shift, u.weekoff,
               a.clock_in, a.clock_out, a.status
        FROM   users u
        LEFT JOIN attendance a ON u.id = a.user_id AND a.date = ?
        WHERE  u.role != 'Admin'
        ORDER  BY u.department, u.username
    ''', (today,)).fetchall()
    conn.close()
    return roster


def get_active_announcements():
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM announcements WHERE active = 1 ORDER BY priority DESC, created_at DESC"
    ).fetchall()
    conn.close()
    return rows


def get_today_stats():
    today = today_str()
    conn  = get_db_connection()
    total   = conn.execute("SELECT COUNT(*) FROM users WHERE role != 'Admin'").fetchone()[0]
    present = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=? AND clock_in IS NOT NULL", (today,)
    ).fetchone()[0]
    absent  = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=? AND status='Absent'", (today,)
    ).fetchone()[0]
    on_leave = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=? AND status='Leave'", (today,)
    ).fetchone()[0]
    late = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=? AND status='Late'", (today,)
    ).fetchone()[0]
    in_office = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=? AND clock_in IS NOT NULL AND clock_out IS NULL", (today,)
    ).fetchone()[0]
    conn.close()
    not_arrived = total - present - absent - on_leave
    return {
        'total': total,
        'present': present,
        'absent': absent,
        'on_leave': on_leave,
        'late': late,
        'in_office': in_office,
        'not_arrived': max(not_arrived, 0),
    }


# ---------------------------------------------------------------------------
# PIN Generation & Verification (NO EMAIL NEEDED!)
# ---------------------------------------------------------------------------
def generate_pin():
    """Generate 4-digit PIN"""
    return ''.join(random.choices(string.digits, k=4))


def log_audit(action, details=None, user_id=None, conn=None):
    """
    Log admin actions.
    If conn is provided, reuse it (same transaction, no lock contention).
    Otherwise open a short-lived connection.
    """
    own_conn = conn is None
    try:
        ip_addr = request.remote_addr if request else 'system'
        if own_conn:
            conn = sqlite3.connect(DB_PATH, timeout=10)
            conn.execute("PRAGMA journal_mode=WAL")
        conn.execute(
            "INSERT INTO audit_logs (user_id, action, details, ip_addr, timestamp) VALUES (?,?,?,?,?)",
            (user_id, action, details, ip_addr, ist_now().strftime('%Y-%m-%d %H:%M:%S'))
        )
        if own_conn:
            conn.commit()
            conn.close()
    except Exception as e:
        logger.error(f"Audit log error: {e}")


# ---------------------------------------------------------------------------
# NOTIFICATION HELPERS
# ---------------------------------------------------------------------------

def push_notification(user_id, title, message, notif_type='info', conn=None):
    """Push an in-app notification to a user"""
    own_conn = conn is None
    try:
        if own_conn:
            conn = get_db_connection()
        conn.execute(
            "INSERT INTO notifications (user_id, title, message, type, created_at) VALUES (?,?,?,?,?)",
            (user_id, title, message, notif_type, ist_now().strftime('%Y-%m-%d %H:%M:%S'))
        )
        if own_conn:
            conn.commit()
            conn.close()
    except Exception as e:
        logger.error(f"Notification push error: {e}")


def get_unread_count(user_id):
    """Get unread notification count for a user"""
    try:
        conn = get_db_connection()
        count = conn.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0",
            (user_id,)
        ).fetchone()[0]
        conn.close()
        return count
    except Exception:
        return 0


def ensure_leave_balance(user_id, conn=None):
    """Ensure leave balance record exists for user"""
    own_conn = conn is None
    try:
        if own_conn:
            conn = get_db_connection()
        existing = conn.execute(
            "SELECT id FROM leave_balance WHERE user_id=?", (user_id,)
        ).fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO leave_balance (user_id, casual, sick, earned) VALUES (?,12,12,15)",
                (user_id,)
            )
            if own_conn:
                conn.commit()
        if own_conn:
            conn.close()
    except Exception as e:
        logger.error(f"Leave balance init error: {e}")


def get_leave_balance(user_id):
    """Get leave balance for a user"""
    try:
        conn = get_db_connection()
        ensure_leave_balance(user_id, conn)
        bal = conn.execute(
            "SELECT casual, sick, earned FROM leave_balance WHERE user_id=?",
            (user_id,)
        ).fetchone()
        conn.close()
        return dict(bal) if bal else {'casual': 12, 'sick': 12, 'earned': 15}
    except Exception:
        return {'casual': 12, 'sick': 12, 'earned': 15}


# ---------------------------------------------------------------------------
# Brute-force protection
# ---------------------------------------------------------------------------
def _check_lockout(username):
    entry = _login_attempts.get(username)
    if not entry:
        return False, 0
    count, locked_until = entry
    if locked_until and ist_now() < locked_until:
        remaining = int((locked_until - ist_now()).total_seconds())
        return True, remaining
    return False, 0


def _record_failed_attempt(username):
    entry = _login_attempts.get(username, [0, None])
    count = entry[0] + 1
    locked_until = None
    if count >= MAX_ATTEMPTS:
        locked_until = ist_now() + timedelta(minutes=LOCKOUT_MINUTES)
        logger.warning("Account locked: %s after %d failed attempts", username, count)
    _login_attempts[username] = [count, locked_until]


def _clear_attempts(username):
    _login_attempts.pop(username, None)


# ---------------------------------------------------------------------------
# Backup Functions
# ---------------------------------------------------------------------------
def create_backup():
    """Create database backup"""
    try:
        timestamp = ist_now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(BACKUP_DIR, f'attendance_backup_{timestamp}.db')
        shutil.copy2(DB_PATH, backup_file)
        logger.info(f"Backup created: {backup_file}")
        return backup_file
    except Exception as e:
        logger.error(f"Backup failed: {e}")
        return None


def get_backups():
    """List all available backups"""
    try:
        backups = []
        for file in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if file.startswith('attendance_backup_') and file.endswith('.db'):
                filepath = os.path.join(BACKUP_DIR, file)
                size = os.path.getsize(filepath)
                mtime = os.path.getmtime(filepath)
                backups.append({
                    'filename': file,
                    'filepath': filepath,
                    'size': size,
                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        return backups[:7]  # Keep last 7 backups
    except Exception as e:
        logger.error(f"Get backups failed: {e}")
        return []


def restore_backup(backup_file):
    """Restore from backup"""
    try:
        shutil.copy2(backup_file, DB_PATH)
        logger.info(f"Database restored from {backup_file}")
        return True
    except Exception as e:
        logger.error(f"Restore failed: {e}")
        return False


def cleanup_old_backups():
    """Keep only last 7 backups"""
    try:
        backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith('attendance_backup_')])
        if len(backups) > 7:
            for old_backup in backups[:-7]:
                os.remove(os.path.join(BACKUP_DIR, old_backup))
    except Exception as e:
        logger.error(f"Cleanup failed: {e}")


# ---------------------------------------------------------------------------
# Context Processor
# ---------------------------------------------------------------------------
@app.context_processor
def inject_globals():
    company_name = "Enterprise"
    try:
        conn  = get_db_connection()
        comp  = conn.execute("SELECT name FROM company LIMIT 1").fetchone()
        conn.close()
        if comp:
            company_name = comp['name']
    except Exception:
        pass
    return dict(
        company_name=company_name,
        current_year=ist_now().year,
    )


# ---------------------------------------------------------------------------
# Before Request
# ---------------------------------------------------------------------------
@app.before_request
def check_setup():
    if request.endpoint in ('setup', 'static'):
        return
    try:
        conn = get_db_connection()
        comp = conn.execute("SELECT * FROM company").fetchone()
        conn.close()
        if not comp:
            return redirect(url_for('setup'))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
@app.route('/setup', methods=['GET', 'POST'])
def setup():
    conn = get_db_connection()
    comp = conn.execute("SELECT * FROM company").fetchone()
    if comp:
        conn.close()
        return redirect(url_for('login'))

    if request.method == 'POST':
        company_name = request.form.get('company_name', '').strip()
        admin_user   = request.form.get('admin_username', '').strip()
        admin_pass   = request.form.get('admin_password', '')
        sec_q        = request.form.get('security_question', '').strip()
        sec_a        = request.form.get('security_answer', '').strip().lower()

        if not all([company_name, admin_user, admin_pass, sec_q, sec_a]):
            flash("All fields are required.")
            conn.close()
            return render_template('setup.html')

        hashed = generate_password_hash(admin_pass)
        conn.execute("INSERT INTO company (name) VALUES (?)", (company_name,))
        existing_admin = conn.execute("SELECT id FROM users WHERE role='Admin'").fetchone()
        if existing_admin:
            conn.execute(
                "UPDATE users SET username=?, password=?, security_question=?, security_answer=? WHERE id=?",
                (admin_user, hashed, sec_q, sec_a, existing_admin['id'])
            )
        else:
            conn.execute(
                "INSERT INTO users (username, password, department, role, shift, weekoff, security_question, security_answer) VALUES (?,?,?,?,?,?,?,?)",
                (admin_user, hashed, 'Management', 'Admin', 'Flexible', 'Sunday', sec_q, sec_a)
            )
        conn.commit()
        conn.close()
        create_backup()  # Create first backup after setup
        flash("✅ System initialized! Welcome to your portal.")
        return redirect(url_for('login'))

    conn.close()
    return render_template('setup.html')


# ---------------------------------------------------------------------------
# NEW LOGIN with PIN (Not Email OTP!)
# ---------------------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        login_type = request.form.get('login_type', 'staff')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        # Check lockout
        locked, secs = _check_lockout(username)
        if locked:
            mins = secs // 60 + 1
            flash(f"Account locked. Try again in {mins} minute(s).")
            return render_template('login.html')

        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()

        # Verify role matches login type
        expected_role = 'Admin' if login_type == 'admin' else 'Staff'
        if user and user['role'] == expected_role and check_password_hash(user['password'], password):
            if user['role'] != 'Admin' and user['department'] not in ALLOWED_DEPARTMENTS:
                flash('Access denied: unauthorized department.')
                return redirect(url_for('login'))

            # Generate PIN (4 digits, NO email needed!)
            pin = generate_pin()
            pin_expires = ist_now() + timedelta(minutes=PIN_EXPIRY_MINUTES)
            _pin_storage[username] = {
                'pin': pin,
                'expires': pin_expires,
                'role': expected_role,
                'user_id': user['id'],
                'department': user['department']
            }

            # Store in session for verification page
            session['pending_login'] = {
                'username': username,
                'role': expected_role,
                'user_id': user['id'],
                'department': user['department']
            }

            flash(f"✅ Your PIN is: {pin}", "success")
            return render_template('pin_verify.html', username=username, pin_display=pin, login_type=login_type)
        else:
            _record_failed_attempt(username)
            count = _login_attempts.get(username, [0])[0]
            remaining = MAX_ATTEMPTS - count
            if remaining > 0:
                flash(f"Invalid credentials. {remaining} attempt(s) left.")
            else:
                flash(f"Account locked for {LOCKOUT_MINUTES} minutes.")

    return render_template('login.html')


@app.route('/verify-pin', methods=['POST'])
def verify_pin():
    """Verify PIN and complete login"""
    pin_entered = request.form.get('pin', '').strip()
    username = request.form.get('username', '').strip()

    if 'pending_login' not in session:
        flash("Session expired. Please login again.")
        return redirect(url_for('login'))
    
    if session.get('pending_login', {}).get('username') != username:
        flash("Session mismatch. Please login again.")
        return redirect(url_for('login'))

    if username not in _pin_storage:
        flash("PIN expired. Please login again.")
        if 'pending_login' in session:
            del session['pending_login']
        return redirect(url_for('login'))

    pin_data = _pin_storage[username]

    # Check PIN expiry
    if ist_now() > pin_data['expires']:
        flash("PIN expired. Please login again.")
        del _pin_storage[username]
        if 'pending_login' in session:
            del session['pending_login']
        return redirect(url_for('login'))

    # Verify PIN
    if pin_data['pin'] != pin_entered:
        flash("Invalid PIN. Try again.")
        # Re-render with the same PIN displayed again
        pending = session.get('pending_login', {})
        login_type = pending.get('role', 'Staff').lower()
        return render_template('pin_verify.html', username=username, pin_display=pin_data['pin'], login_type=login_type)

    # PIN verified! Complete login
    pending = session.get('pending_login', {})
    if not pending:
        flash("Session expired. Please login again.")
        return redirect(url_for('login'))
    
    _clear_attempts(username)
    del _pin_storage[username]
    del session['pending_login']

    session['user_id']    = pending.get('user_id')
    session['username']   = username
    session['department'] = pending.get('department')
    session['role']       = pending.get('role')

    # Store display name (full_name > username)
    try:
        conn = get_db_connection()
        urow = conn.execute("SELECT full_name FROM users WHERE id=?", (pending.get('user_id'),)).fetchone()
        conn.close()
        if urow:
            session['display_name'] = urow['full_name'] or username
        else:
            session['display_name'] = username
    except Exception:
        session['display_name'] = username

    log_audit('LOGIN_SUCCESS', f"Role: {pending.get('role')}", pending.get('user_id'))
    logger.info("Login successful: %s (%s)", username, pending.get('role'))
    flash("✅ Logged in successfully!", "success")

    # Staff → trigger agent download page first
    if pending.get('role') == 'Staff':
        return redirect(url_for('agent_launch'))

    return redirect(url_for('index'))


@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    username = session.get('username')
    role = session.get('role')
    log_audit('LOGOUT', f"Role: {role}", user_id)
    logger.info("Logout: %s", username)
    session.clear()
    flash("Logged out successfully.")
    return redirect(url_for('login'))


@app.route('/forgot', methods=['GET', 'POST'])
def forgot():
    step = request.args.get('step', '1')
    if request.method == 'POST':
        conn = get_db_connection()
        if step == '1':
            username = request.form.get('username', '').strip()
            user = conn.execute(
                "SELECT security_question FROM users WHERE username = ?", (username,)
            ).fetchone()
            conn.close()
            if user:
                return render_template('forgot.html', step='2', username=username, question=user['security_question'])
            flash("No account found with that username.")
            return redirect(url_for('forgot'))

        elif step == '2':
            username = request.form.get('username', '').strip()
            answer   = request.form.get('security_answer', '').strip().lower()
            new_pass = request.form.get('new_password', '')

            if not new_pass or len(new_pass) < 6:
                conn.close()
                flash("New password must be at least 6 characters.")
                return redirect(url_for('forgot'))

            user = conn.execute(
                "SELECT id, security_answer FROM users WHERE username = ?", (username,)
            ).fetchone()
            if user and user['security_answer'] == answer:
                conn.execute(
                    "UPDATE users SET password = ? WHERE username = ?",
                    (generate_password_hash(new_pass), username)
                )
                conn.commit()
                conn.close()
                _clear_attempts(username)
                log_audit('PASSWORD_RESET', str(user['id']), user['id'])
                flash("✅ Password reset successfully. You may now log in.")
                return redirect(url_for('login'))
            else:
                conn.close()
                flash("Incorrect security answer. Access denied.")
                return redirect(url_for('forgot'))

    return render_template('forgot.html', step='1')


# ---------------------------------------------------------------------------
# Index & Routing
# ---------------------------------------------------------------------------
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session['role'] == 'Admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('staff_dashboard'))


# ---------------------------------------------------------------------------
# Staff Dashboard
# ---------------------------------------------------------------------------
@app.route('/staff')
def staff_dashboard():
    if 'user_id' not in session or session['role'] != 'Staff':
        return redirect(url_for('login'))

    today_day = ist_now().strftime('%A')
    conn      = get_db_connection()
    user_data = conn.execute(
        'SELECT shift, weekoff FROM users WHERE id = ?', (session['user_id'],)
    ).fetchone()
    logs = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? ORDER BY date DESC LIMIT 30',
        (session['user_id'],)
    ).fetchall()
    conn.close()

    return render_template(
        'staff.html',
        logs=logs,
        roster=get_todays_roster(),
        today_day=today_day,
        user_shift=user_data['shift'] if user_data else '09:00 AM - 06:00 PM',
        user_weekoff=user_data['weekoff'] if user_data else 'Sunday',
        announcements=get_active_announcements(),
    )


@app.route('/staff/history')
def staff_history():
    if 'user_id' not in session or session['role'] != 'Staff':
        return redirect(url_for('login'))
    conn = get_db_connection()
    logs = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? ORDER BY date DESC',
        (session['user_id'],)
    ).fetchall()
    conn.close()
    return render_template('history.html', logs=logs)


@app.route('/clock', methods=['POST'])
def clock():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    action   = request.form.get('action')
    user_id  = session['user_id']
    now      = ist_now()
    today    = now.strftime('%Y-%m-%d')
    now_time = now.strftime('%H:%M:%S')

    conn   = get_db_connection()
    record = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? AND date = ?', (user_id, today)
    ).fetchone()

    if action == 'in':
        if not record:
            conn.execute(
                'INSERT INTO attendance (user_id, date, clock_in, status) VALUES (?,?,?,?)',
                (user_id, today, now_time, 'Present')
            )
            log_audit('CLOCK_IN', today, user_id, conn=conn)
            flash(f'✅ Clocked in at {now_time}')
        elif not record['clock_in']:
            # Row already exists but clock_in is empty — this happens when the
            # background scheduler auto-marked the day Late/Absent before the
            # staff member got a chance to clock in. Let them clock in now
            # instead of being permanently locked out for the day.
            new_status = 'Present' if record['status'] == 'Absent' else record['status']
            conn.execute(
                'UPDATE attendance SET clock_in = ?, status = ? WHERE id = ?',
                (now_time, new_status, record['id'])
            )
            log_audit('CLOCK_IN', today, user_id, conn=conn)
            flash(f'✅ Clocked in at {now_time}')
        else:
            flash('You have already clocked in today.')
    elif action == 'out':
        if record and not record['clock_out']:
            conn.execute(
                'UPDATE attendance SET clock_out = ? WHERE id = ?', (now_time, record['id'])
            )
            log_audit('CLOCK_OUT', today, user_id, conn=conn)
            flash(f'👋 Clocked out at {now_time}')
        else:
            flash('You must clock in first, or have already clocked out.')

    conn.commit()
    conn.close()
    return redirect(url_for('staff_dashboard'))


# ---------------------------------------------------------------------------
# Admin Dashboard
# ---------------------------------------------------------------------------
@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    today_day = ist_now().strftime('%A')
    conn  = get_db_connection()
    users = conn.execute(
        "SELECT id, username, full_name, department, shift, weekoff FROM users WHERE role != 'Admin' ORDER BY department, username"
    ).fetchall()
    announcements = conn.execute(
        "SELECT * FROM announcements ORDER BY created_at DESC LIMIT 20"
    ).fetchall()
    conn.close()

    return render_template(
        'admin.html',
        users=users,
        roster=get_todays_roster(),
        today_day=today_day,
        shift_options=SHIFT_OPTIONS,
        weekday_options=WEEKDAY_OPTIONS,
        departments=ALLOWED_DEPARTMENTS,
        stats=get_today_stats(),
        announcements=announcements,
    )


@app.route('/admin/analytics')
def admin_analytics():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    conn = get_db_connection()
    rows = conn.execute('''
        SELECT
            a.date,
            COUNT(DISTINCT a.user_id) AS total_marked,
            SUM(CASE WHEN a.clock_in IS NOT NULL THEN 1 ELSE 0 END) AS present,
            SUM(CASE WHEN a.status = 'Absent' THEN 1 ELSE 0 END) AS absent,
            SUM(CASE WHEN a.status = 'Leave' THEN 1 ELSE 0 END) AS on_leave
        FROM attendance a
        WHERE a.date >= date('now', '-30 days')
        GROUP BY a.date
        ORDER BY a.date DESC
    ''').fetchall()

    user_summary = conn.execute('''
        SELECT
            u.username, u.department,
            COUNT(a.id) AS total_days,
            SUM(CASE WHEN a.clock_in IS NOT NULL THEN 1 ELSE 0 END) AS present_days,
            SUM(CASE WHEN a.status = 'Absent' THEN 1 ELSE 0 END) AS absent_days,
            SUM(CASE WHEN a.status = 'Leave' THEN 1 ELSE 0 END) AS leave_days
        FROM users u
        LEFT JOIN attendance a ON u.id = a.user_id AND a.date >= date('now', '-30 days')
        WHERE u.role != 'Admin'
        GROUP BY u.id
        ORDER BY absent_days DESC, u.username
    ''').fetchall()

    conn.close()
    return render_template('analytics.html', daily_rows=rows, user_summary=user_summary)


@app.route('/admin/system')
def admin_system():
    """Server status & backup management"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    try:
        db_size = os.path.getsize(DB_PATH) / (1024 * 1024)  # MB
    except:
        db_size = 0

    backups = get_backups()
    conn = get_db_connection()
    user_count = conn.execute("SELECT COUNT(*) FROM users WHERE role != 'Admin'").fetchone()[0]
    conn.close()

    return render_template('system.html', db_size=db_size, backups=backups, user_count=user_count)


@app.route('/admin/backup', methods=['POST'])
def admin_backup():
    """Create backup manually"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    backup_file = create_backup()
    if backup_file:
        log_audit('BACKUP_CREATED', backup_file, session['user_id'])
        flash(f"✅ Backup created successfully!")
    else:
        flash("❌ Backup failed. Check disk space.")

    return redirect(url_for('admin_system'))


@app.route('/admin/restore/<backup_name>', methods=['POST'])
def admin_restore(backup_name):
    """Restore from backup"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    backup_file = os.path.join(BACKUP_DIR, backup_name)
    
    # Security check: ensure file is in backup directory
    if not backup_file.startswith(BACKUP_DIR) or not os.path.exists(backup_file):
        flash("❌ Invalid backup file.")
        return redirect(url_for('admin_system'))

    if restore_backup(backup_file):
        log_audit('DATABASE_RESTORED', backup_name, session['user_id'])
        flash(f"✅ Database restored from {backup_name}!")
    else:
        flash("❌ Restore failed. Check permissions.")

    return redirect(url_for('admin_system'))


@app.route('/admin_action', methods=['POST'])
def admin_action():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    action_type = request.form.get('action_type')
    conn = get_db_connection()

    try:
        if action_type == 'add_user':
            new_user  = request.form.get('new_username', '').strip()
            new_pass  = request.form.get('new_password', '')
            full_name = request.form.get('full_name', '').strip()
            dept      = request.form.get('department', '')
            weekoff   = request.form.get('weekoff', 'Sunday')

            if not new_user or not new_pass or dept not in ALLOWED_DEPARTMENTS:
                flash("Invalid input.")
            elif weekoff not in WEEKDAY_OPTIONS:
                flash("Invalid week-off day.")
            else:
                try:
                    conn.execute(
                        "INSERT INTO users (username, password, full_name, department, role, shift, weekoff, security_question, security_answer) VALUES (?,?,?,?,?,?,?,?,?)",
                        (new_user, generate_password_hash(new_pass), full_name,
                         dept, 'Staff', '09:00 AM - 06:00 PM', weekoff, 'Set by admin', 'yes')
                    )
                    log_audit('USER_CREATED', f"{new_user} ({dept})", session['user_id'], conn=conn)
                    flash(f"✅ Employee '{new_user}' created.")
                except sqlite3.IntegrityError:
                    flash(f"Username '{new_user}' already exists.")

        elif action_type == 'delete_user':
            target = request.form.get('target_user', '').strip()
            user = conn.execute("SELECT id FROM users WHERE username=? AND role!='Admin'", (target,)).fetchone()
            if user:
                conn.execute("DELETE FROM users WHERE id = ?", (user['id'],))
                log_audit('USER_DELETED', target, session['user_id'], conn=conn)
                flash(f"✅ Employee '{target}' removed.")
            else:
                flash("User not found.")

        elif action_type == 'reset_password':
            target   = request.form.get('target_user', '').strip()
            new_pass = request.form.get('new_password', '')
            if len(new_pass) < 6:
                flash("Password must be 6+ characters.")
            else:
                conn.execute(
                    "UPDATE users SET password=? WHERE username=?",
                    (generate_password_hash(new_pass), target)
                )
                log_audit('PASSWORD_RESET_ADMIN', target, session['user_id'], conn=conn)
                flash(f"✅ Password reset for '{target}'.")

        elif action_type == 'change_shift':
            target      = request.form.get('target_user', '').strip()
            new_shift   = request.form.get('new_shift', '')
            new_weekoff = request.form.get('new_weekoff', '')
            if new_shift not in SHIFT_OPTIONS or new_weekoff not in WEEKDAY_OPTIONS:
                flash("Invalid shift or week-off.")
            else:
                conn.execute(
                    "UPDATE users SET shift=?, weekoff=? WHERE username=?",
                    (new_shift, new_weekoff, target)
                )
                log_audit('SHIFT_UPDATED', f"{target}: {new_shift}", session['user_id'], conn=conn)
                flash(f"✅ Schedule updated for '{target}'.")

        elif action_type == 'mark_leave':
            target = request.form.get('target_user', '').strip()
            status = request.form.get('status', '')
            if status not in ('Leave', 'Absent'):
                flash("Invalid status.")
            else:
                today = today_str()
                user  = conn.execute('SELECT id FROM users WHERE username=?', (target,)).fetchone()
                if user:
                    record = conn.execute(
                        'SELECT id FROM attendance WHERE user_id=? AND date=?', (user['id'], today)
                    ).fetchone()
                    if record:
                        conn.execute('UPDATE attendance SET status=? WHERE id=?', (status, record['id']))
                    else:
                        conn.execute(
                            'INSERT INTO attendance (user_id, date, status) VALUES (?,?,?)',
                            (user['id'], today, status)
                        )
                    log_audit('STATUS_MARKED', f"{target}: {status}", session['user_id'], conn=conn)
                    flash(f"✅ '{target}' marked as {status}.")
                else:
                    flash("User not found.")

        elif action_type == 'post_announcement':
            title    = request.form.get('ann_title', '').strip()
            body     = request.form.get('ann_body', '').strip()
            priority = request.form.get('ann_priority', 'normal')
            if not title or not body:
                flash("Title and message required.")
            elif priority not in ('normal', 'high', 'urgent'):
                flash("Invalid priority.")
            else:
                conn.execute(
                    "INSERT INTO announcements (title, body, priority, created_at, created_by, active) VALUES (?,?,?,?,?,1)",
                    (title, body, priority, ist_now().strftime('%Y-%m-%d %H:%M:%S'), session['username'])
                )
                log_audit('ANNOUNCEMENT_POSTED', title, session['user_id'], conn=conn)
                flash(f"✅ Announcement posted.")

        elif action_type == 'delete_announcement':
            ann_id = request.form.get('ann_id', '')
            if ann_id.isdigit():
                conn.execute("DELETE FROM announcements WHERE id=?", (int(ann_id),))
                log_audit('ANNOUNCEMENT_DELETED', f"ID: {ann_id}", session['user_id'], conn=conn)
                flash("✅ Announcement removed.")
            else:
                flash("Invalid ID.")

        elif action_type == 'update_profile':
            target    = request.form.get('target_user', '').strip()
            full_name = request.form.get('full_name', '').strip()
            conn.execute(
                "UPDATE users SET full_name=? WHERE username=?",
                (full_name, target)
            )
            log_audit('PROFILE_UPDATED', f"{target}: {full_name}", session['user_id'], conn=conn)
            flash(f"✅ Profile updated for '{target}'.")

        elif action_type == 'update_company':
            new_name = request.form.get('company_name', '').strip()
            if not new_name:
                flash("Company name required.")
            else:
                conn.execute("UPDATE company SET name=? WHERE id=1", (new_name,))
                log_audit('COMPANY_UPDATED', new_name, session['user_id'], conn=conn)
                flash(f"✅ Company name updated.")

        conn.commit()

    except Exception as e:
        conn.rollback()
        logger.error(f"Admin action error: {e}")
        flash("❌ Error. Please try again.")
    finally:
        conn.close()

    return redirect(url_for('admin_dashboard'))


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
@app.route('/export/<string:report_type>')
def export_excel(report_type):
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    if report_type not in ('weekly', 'monthly', 'all'):
        flash("Invalid report type.")
        return redirect(url_for('admin_dashboard'))

    target_user = request.args.get('target_user', 'All')
    conn = get_db_connection()
    df   = pd.read_sql_query('''
        SELECT u.username, u.full_name, u.department, u.shift, u.weekoff,
               a.date, a.clock_in, a.clock_out, a.status
        FROM   attendance a
        JOIN   users u ON a.user_id = u.id
        ORDER  BY a.date DESC
    ''', conn)
    conn.close()

    if df.empty:
        flash("No attendance data.")
        return redirect(url_for('admin_dashboard'))

    df['date'] = pd.to_datetime(df['date'])
    now = datetime.now()
    if report_type == 'weekly':
        df = df[df['date'] >= (now - timedelta(days=7))]
    elif report_type == 'monthly':
        df = df[df['date'] >= (now - timedelta(days=30))]

    if target_user != 'All':
        df = df[df['username'] == target_user]

    if df.empty:
        flash("No records found.")
        return redirect(url_for('admin_dashboard'))

    df['date']  = df['date'].dt.strftime('%Y-%m-%d')
    safe_user   = target_user.replace(' ', '_')
    output_path = f"/tmp/{report_type}_attendance_{safe_user}.xlsx"
    df.to_excel(output_path, index=False)
    log_audit('EXPORT', f"{report_type} ({target_user})", session['user_id'])
    return send_file(output_path, as_attachment=True)


# ---------------------------------------------------------------------------
# Monthly Attendance Register — replaces the manual paper register.
# One sheet per employee status grid (day-by-day, like a manual muster roll)
# plus a Summary sheet with totals per employee for the month.
# ---------------------------------------------------------------------------
STATUS_CODE = {
    'Present': 'P',
    'Late':    'L',
    'Absent':  'A',
    'Leave':   'LV',
}
STATUS_FILL = {
    'P':  PatternFill('solid', fgColor='D8F3DC'),
    'L':  PatternFill('solid', fgColor='FFF3CD'),
    'A':  PatternFill('solid', fgColor='F8D7DA'),
    'LV': PatternFill('solid', fgColor='D6E4FF'),
    'WO': PatternFill('solid', fgColor='E9ECEF'),
    '-':  PatternFill('solid', fgColor='FFFFFF'),
}
THIN_BORDER = Border(*(Side(style='thin', color='D0D3D9'),) * 4)


@app.route('/reports/register')
def attendance_register():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    try:
        ym = request.args.get('ym')
        if ym and '-' in ym:
            year, month = (int(p) for p in ym.split('-', 1))
        else:
            year  = int(request.args.get('year', ist_now().year))
            month = int(request.args.get('month', ist_now().month))
        if not (1 <= month <= 12) or not (2000 <= year <= 2100):
            raise ValueError
    except (TypeError, ValueError):
        flash("Invalid month/year.")
        return redirect(url_for('admin_dashboard'))

    conn  = get_db_connection()
    users = conn.execute(
        "SELECT id, username, full_name, department, weekoff FROM users "
        "WHERE role='Staff' ORDER BY department, username"
    ).fetchall()

    if not users:
        conn.close()
        flash("No staff to report on.")
        return redirect(url_for('admin_dashboard'))

    days_in_month = calendar.monthrange(year, month)[1]
    today         = ist_now().date()
    month_start   = datetime(year, month, 1).date()

    # Pull the whole month's records in one query, keyed by (user_id, date)
    month_str = f"{year:04d}-{month:02d}"
    records = conn.execute(
        "SELECT user_id, date, status FROM attendance WHERE date LIKE ?",
        (f"{month_str}-%",)
    ).fetchall()
    conn.close()
    status_by_user_date = {(r['user_id'], r['date']): r['status'] for r in records}

    wb = Workbook()

    # ---- Register sheet (day-by-day grid) ----
    ws = wb.active
    ws.title = "Register"
    header_fill = PatternFill('solid', fgColor='343A40')
    header_font = Font(color='FFFFFF', bold=True)

    ws.cell(row=1, column=1, value="Employee").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value="Department").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    for d in range(1, days_in_month + 1):
        c = ws.cell(row=1, column=2 + d, value=d)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for i, u in enumerate(users):
        row = i + 2
        ws.cell(row=row, column=1, value=u['full_name'] or u['username'])
        ws.cell(row=row, column=2, value=u['department'] or '-')
        for d in range(1, days_in_month + 1):
            date_obj = datetime(year, month, d).date()
            date_str = date_obj.strftime('%Y-%m-%d')
            weekday_name = date_obj.strftime('%A')

            if date_obj > today:
                code = ''
            elif weekday_name == u['weekoff']:
                code = 'WO'
            else:
                status = status_by_user_date.get((u['id'], date_str))
                code = STATUS_CODE.get(status, 'A' if date_obj < today else '-')

            cell = ws.cell(row=row, column=2 + d, value=code)
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            if code in STATUS_FILL:
                cell.fill = STATUS_FILL[code]

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(2 + d)].width = 4
    ws.freeze_panes = "C2"

    legend_row = len(users) + 4
    ws.cell(row=legend_row, column=1, value="Legend: P=Present  L=Late  A=Absent  LV=Leave  WO=Week Off  (blank)=Future date").font = Font(italic=True, size=9)

    # ---- Summary sheet (totals per employee) ----
    ws2 = wb.create_sheet("Summary")
    headers = ["Employee", "Department", "Present", "Late", "Absent", "Leave", "Week Off", "Days Marked", "Days in Month"]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, u in enumerate(users):
        row = i + 2
        counts = {'P': 0, 'L': 0, 'A': 0, 'LV': 0, 'WO': 0}
        for d in range(1, days_in_month + 1):
            date_obj = datetime(year, month, d).date()
            if date_obj > today:
                continue
            date_str = date_obj.strftime('%Y-%m-%d')
            weekday_name = date_obj.strftime('%A')
            if weekday_name == u['weekoff']:
                counts['WO'] += 1
            else:
                status = status_by_user_date.get((u['id'], date_str))
                code = STATUS_CODE.get(status, 'A')
                counts[code] = counts.get(code, 0) + 1

        days_marked = counts['P'] + counts['L'] + counts['A'] + counts['LV']
        ws2.cell(row=row, column=1, value=u['full_name'] or u['username'])
        ws2.cell(row=row, column=2, value=u['department'] or '-')
        ws2.cell(row=row, column=3, value=counts['P'])
        ws2.cell(row=row, column=4, value=counts['L'])
        ws2.cell(row=row, column=5, value=counts['A'])
        ws2.cell(row=row, column=6, value=counts['LV'])
        ws2.cell(row=row, column=7, value=counts['WO'])
        ws2.cell(row=row, column=8, value=days_marked)
        ws2.cell(row=row, column=9, value=days_in_month)

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    month_name = calendar.month_name[month]
    output_path = f"/tmp/attendance_register_{year}_{month:02d}.xlsx"
    wb.save(output_path)
    log_audit('EXPORT_REGISTER', f"{month_name} {year}", session['user_id'])
    return send_file(
        output_path,
        as_attachment=True,
        download_name=f"Attendance_Register_{month_name}_{year}.xlsx"
    )


# ---------------------------------------------------------------------------
# Bulk User Upload — add employees in bulk from an Excel or PDF file.
# Shows the required format / a downloadable sample before uploading.
# ---------------------------------------------------------------------------
BULK_UPLOAD_COLUMNS = [
    # (canonical key, display header, required, help text)
    ('username',   'Username',   True,  'Unique login ID, no spaces'),
    ('full_name',  'Full Name',  False, "Defaults to username if left blank"),
    ('password',   'Password',   False, 'Auto-generated if left blank'),
    ('department', 'Department', True,  ' / '.join(ALLOWED_DEPARTMENTS)),
    ('weekoff',    'Weekoff',    False, "Default: Sunday. One of: " + ', '.join(WEEKDAY_OPTIONS)),
    ('shift',      'Shift',      False, "Default: 09:00 AM - 06:00 PM. One of: " + ' / '.join(SHIFT_OPTIONS)),
]
BULK_SAMPLE_ROWS = [
    ['jdoe',  'John Doe',  '',           'IT', 'Sunday',   '09:00 AM - 06:00 PM'],
    ['asmith','Amy Smith', 'Amy@12345',  'QA', 'Saturday', '10:00 AM - 07:00 PM'],
]
_COLUMN_ALIASES = {
    'username': 'username', 'user name': 'username', 'user id': 'username', 'userid': 'username',
    'full name': 'full_name', 'fullname': 'full_name', 'name': 'full_name',
    'password': 'password', 'pass': 'password',
    'department': 'department', 'dept': 'department',
    'weekoff': 'weekoff', 'week off': 'weekoff', 'week-off': 'weekoff',
    'shift': 'shift',
}


def _normalize_col(col):
    key = str(col).strip().lower()
    return _COLUMN_ALIASES.get(key, key.replace(' ', '_'))


def _generate_temp_password():
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(10))


def _build_sample_workbook():
    """In-memory .xlsx sample template with headers, example rows, and
    dropdown validation for Department / Weekoff / Shift."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    header_fill = PatternFill('solid', fgColor='343A40')
    header_font = Font(color='FFFFFF', bold=True)

    headers = [h for _, h, _, _ in BULK_UPLOAD_COLUMNS]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    for r, row in enumerate(BULK_SAMPLE_ROWS, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Dropdown validation so admins can't typo department/weekoff/shift
    dept_dv = DataValidation(type="list", formula1=f'"{",".join(ALLOWED_DEPARTMENTS)}"', allow_blank=False)
    weekoff_dv = DataValidation(type="list", formula1=f'"{",".join(WEEKDAY_OPTIONS)}"', allow_blank=True)
    shift_dv = DataValidation(type="list", formula1=f'"{",".join(SHIFT_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(dept_dv)
    ws.add_data_validation(weekoff_dv)
    ws.add_data_validation(shift_dv)
    dept_dv.add(f"D2:D200")
    weekoff_dv.add(f"E2:E200")
    shift_dv.add(f"F2:F200")

    notes = wb.create_sheet("Instructions")
    notes.column_dimensions['A'].width = 100
    lines = [
        "Bulk Employee Upload — Instructions",
        "",
        "Fill one row per employee on the 'Employees' sheet, keeping the header row as-is.",
        "",
    ]
    for key, header, required, help_text in BULK_UPLOAD_COLUMNS:
        req = "REQUIRED" if required else "optional"
        lines.append(f"• {header} ({req}): {help_text}")
    lines += [
        "",
        "Usernames must be unique. Rows with a username that already exists will be skipped.",
        "If Password is left blank, a random temporary password is generated and shown after upload.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = notes.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_sample_pdf():
    """In-memory sample PDF showing the same table format for admins who
    prefer to prepare/upload a PDF instead of an Excel file."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), title="Bulk Employee Upload Sample")
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Bulk Employee Upload — Sample Format", styles['Title']),
        Spacer(1, 10),
        Paragraph(
            "Keep the header row exactly as shown. One row per employee. "
            "Required columns: Username, Department.",
            styles['Normal']
        ),
        Spacer(1, 14),
    ]

    headers = [h for _, h, _, _ in BULK_UPLOAD_COLUMNS]
    data = [headers] + BULK_SAMPLE_ROWS
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D3D9')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F6F8')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 16))

    for key, header, required, help_text in BULK_UPLOAD_COLUMNS:
        req = "Required" if required else "Optional"
        story.append(Paragraph(f"<b>{header}</b> ({req}): {help_text}", styles['Normal']))
        story.append(Spacer(1, 4))

    doc.build(story)
    buf.seek(0)
    return buf


def _parse_bulk_rows(file_storage):
    """Returns (rows: list[dict], error: str|None). Supports .xlsx/.xls/.csv/.pdf"""
    filename = secure_filename(file_storage.filename or '')
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext in ('xlsx', 'xls'):
        df = pd.read_excel(file_storage, sheet_name=0, dtype=str)
    elif ext == 'csv':
        df = pd.read_csv(file_storage, dtype=str)
    elif ext == 'pdf':
        table_rows = None
        with pdfplumber.open(file_storage) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    if table and len(table) > 1:
                        table_rows = table
                        break
                if table_rows:
                    break
        if not table_rows:
            return None, ("Couldn't find a table in that PDF. Make sure it has clear "
                           "column/row lines like the sample PDF, or use the Excel template instead.")
        df = pd.DataFrame(table_rows[1:], columns=table_rows[0])
    else:
        return None, "Unsupported file type. Please upload a .xlsx, .csv, or .pdf file."

    df = df.dropna(how='all')
    df.columns = [_normalize_col(c) for c in df.columns]

    rows = []
    for _, r in df.iterrows():
        row = {key: (str(r[key]).strip() if key in df.columns and pd.notna(r.get(key)) else '')
               for key, _, _, _ in BULK_UPLOAD_COLUMNS}
        if any(row.values()):
            rows.append(row)
    return rows, None


@app.route('/admin/bulk-upload', methods=['GET', 'POST'])
def bulk_upload_users():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    if request.method == 'GET':
        return render_template('bulk_upload.html', columns=BULK_UPLOAD_COLUMNS, sample_rows=BULK_SAMPLE_ROWS)

    uploaded = request.files.get('bulk_file')
    if not uploaded or not uploaded.filename:
        flash("Please choose a file to upload.")
        return redirect(url_for('bulk_upload_users'))

    rows, error = _parse_bulk_rows(uploaded)
    if error:
        flash(error)
        return redirect(url_for('bulk_upload_users'))
    if not rows:
        flash("No employee rows found in that file.")
        return redirect(url_for('bulk_upload_users'))

    conn = get_db_connection()
    results = []
    for i, row in enumerate(rows, start=1):
        username = row['username']
        department = row['department']
        full_name = row['full_name'] or username
        weekoff = row['weekoff'] or 'Sunday'
        shift = row['shift'] or SHIFT_OPTIONS[0]
        temp_password = None

        if not username:
            results.append({'row': i, 'username': '(blank)', 'status': 'skipped', 'reason': 'Missing username'})
            continue
        if department not in ALLOWED_DEPARTMENTS:
            results.append({'row': i, 'username': username, 'status': 'skipped',
                             'reason': f"Invalid department '{department}'"})
            continue
        if weekoff not in WEEKDAY_OPTIONS:
            results.append({'row': i, 'username': username, 'status': 'skipped',
                             'reason': f"Invalid weekoff '{weekoff}'"})
            continue
        if shift not in SHIFT_OPTIONS:
            shift = SHIFT_OPTIONS[0]

        password = row['password']
        if not password:
            password = _generate_temp_password()
            temp_password = password

        try:
            conn.execute(
                "INSERT INTO users (username, password, full_name, department, role, shift, weekoff, "
                "security_question, security_answer) VALUES (?,?,?,?,?,?,?,?,?)",
                (username, generate_password_hash(password), full_name, department,
                 'Staff', shift, weekoff, 'Set by admin', 'yes')
            )
            log_audit('USER_CREATED', f"{username} ({department}) [bulk]", session['user_id'], conn=conn)
            results.append({'row': i, 'username': username, 'status': 'created',
                             'reason': f"Temp password: {temp_password}" if temp_password else ''})
        except sqlite3.IntegrityError:
            results.append({'row': i, 'username': username, 'status': 'skipped',
                             'reason': 'Username already exists'})

    conn.commit()
    conn.close()

    created = sum(1 for r in results if r['status'] == 'created')
    skipped = sum(1 for r in results if r['status'] == 'skipped')
    flash(f"✅ Bulk upload complete: {created} created, {skipped} skipped.")
    return render_template('bulk_upload_result.html', results=results, created=created, skipped=skipped)


@app.route('/admin/bulk-upload/sample.xlsx')
def bulk_upload_sample_xlsx():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))
    return send_file(
        _build_sample_workbook(),
        as_attachment=True,
        download_name="employee_upload_sample.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/admin/bulk-upload/sample.pdf')
def bulk_upload_sample_pdf():
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))
    return send_file(
        _build_sample_pdf(),
        as_attachment=True,
        download_name="employee_upload_sample.pdf",
        mimetype="application/pdf"
    )


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# LIVE ACTIVITY TRACKING
# ---------------------------------------------------------------------------

IDLE_SECONDS    = 60    # 1 minute no activity → idle
OFFLINE_SECONDS = 300   # 5 minutes no heartbeat → offline

@app.route('/api/agent-login', methods=['POST'])
def agent_login():
    """Windows agent login — returns session for heartbeat"""
    data     = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({'ok': False, 'error': 'Missing credentials'}), 400

    # Check lockout
    locked, secs = _check_lockout(username)
    if locked:
        return jsonify({'ok': False, 'error': f'Locked for {secs}s'}), 429

    conn = get_db_connection()
    user = conn.execute(
        'SELECT * FROM users WHERE username = ?', (username,)
    ).fetchone()
    conn.close()

    if user and check_password_hash(user['password'], password):
        _clear_attempts(username)
        session['user_id']    = user['id']
        session['username']   = user['username']
        session['department'] = user['department']
        session['role']       = user['role']
        log_audit('AGENT_LOGIN', f"Windows agent login: {username}", user['id'])
        return jsonify({'ok': True, 'role': user['role']})

    _record_failed_attempt(username)
    return jsonify({'ok': False, 'error': 'Invalid credentials'}), 401


@app.route('/api/agent-logout', methods=['POST'])
def agent_logout():
    """Mark agent as offline on shutdown"""
    if 'user_id' not in session:
        return jsonify({'ok': False}), 401

    user_id  = session['user_id']
    username = session['username']
    now_str  = ist_now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        conn = sqlite3.connect(DB_PATH, timeout=5)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute(
            "UPDATE activity_logs SET status='offline', last_seen=? WHERE user_id=?",
            (now_str, user_id)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Agent logout error: {e}")

    log_audit('AGENT_LOGOUT', f"Windows agent logout: {username}", user_id)
    return jsonify({'ok': True})


@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    """Called every 10s from staff browser — silent tracking"""
    if 'user_id' not in session:
        return jsonify({'ok': False}), 401

    data        = request.get_json(silent=True) or {}
    user_id     = session['user_id']
    username    = session['username']
    current_page= data.get('page', '/')
    has_activity= data.get('active', False)  # True = mouse/keyboard detected
    is_offline  = data.get('offline', False)  # True = agent shutting down
    ip_addr     = request.remote_addr
    now_str     = ist_now().strftime('%Y-%m-%d %H:%M:%S')

    # If agent is shutting down, mark offline immediately
    if is_offline:
        try:
            conn = sqlite3.connect(DB_PATH, timeout=5)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute(
                "UPDATE activity_logs SET status='offline', last_seen=? WHERE user_id=?",
                (now_str, user_id)
            )
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Offline mark error: {e}")
        return jsonify({'ok': True})

    try:
        conn = sqlite3.connect(DB_PATH, timeout=5)
        conn.execute("PRAGMA journal_mode=WAL")
        existing = conn.execute(
            "SELECT id, last_active FROM activity_logs WHERE user_id=?", (user_id,)
        ).fetchone()

        if existing:
            # Update last_seen always; update last_active only if activity
            if has_activity:
                conn.execute(
                    "UPDATE activity_logs SET status='active', last_seen=?, last_active=?, current_page=?, ip_addr=? WHERE user_id=?",
                    (now_str, now_str, current_page, ip_addr, user_id)
                )
            else:
                # Compute idle vs active from last_active
                last_active = existing[1] or now_str
                try:
                    la = datetime.strptime(last_active, '%Y-%m-%d %H:%M:%S')
                    la = pytz.timezone('Asia/Kolkata').localize(la)
                    diff = (ist_now() - la).total_seconds()
                except:
                    diff = 0
                status = 'active' if diff < IDLE_SECONDS else 'idle'
                conn.execute(
                    "UPDATE activity_logs SET status=?, last_seen=?, current_page=?, ip_addr=? WHERE user_id=?",
                    (status, now_str, current_page, ip_addr, user_id)
                )
        else:
            conn.execute(
                "INSERT INTO activity_logs (user_id, username, status, current_page, ip_addr, last_seen, last_active) VALUES (?,?,?,?,?,?,?)",
                (user_id, username, 'active', current_page, ip_addr, now_str, now_str)
            )

        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Heartbeat error: {e}")

    return jsonify({'ok': True})


@app.route('/api/activity')
def get_activity():
    """Admin only — returns live activity of all staff"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 401

    now = ist_now()
    try:
        conn = get_db_connection()

        # Get all staff
        staff = conn.execute(
            "SELECT id, username, full_name, department, shift FROM users WHERE role != 'Admin' ORDER BY department, username"
        ).fetchall()

        # Get activity logs
        activity = conn.execute(
            "SELECT user_id, status, current_page, ip_addr, last_seen, last_active FROM activity_logs"
        ).fetchall()
        conn.close()

        activity_map = {row['user_id']: row for row in activity}

        result = []
        for s in staff:
            act = activity_map.get(s['id'])
            if act:
                # Recalculate status based on time elapsed
                last_seen = act['last_seen']
                try:
                    ls = datetime.strptime(last_seen, '%Y-%m-%d %H:%M:%S')
                    ls = pytz.timezone('Asia/Kolkata').localize(ls)
                    elapsed = (now - ls).total_seconds()
                except:
                    elapsed = 9999

                if elapsed > OFFLINE_SECONDS:
                    status = 'offline'
                elif elapsed > IDLE_SECONDS:
                    status = 'idle'
                else:
                    status = act['status'] or 'active'

                last_active = act['last_active'] or '—'
                # Format last seen nicely
                try:
                    ls_dt = datetime.strptime(last_seen, '%Y-%m-%d %H:%M:%S')
                    ls_dt = pytz.timezone('Asia/Kolkata').localize(ls_dt)
                    secs  = int((now - ls_dt).total_seconds())
                    if secs < 60:   ago = f"{secs}s ago"
                    elif secs < 3600: ago = f"{secs//60}m ago"
                    else:           ago = f"{secs//3600}h ago"
                except:
                    ago = '—'

                result.append({
                    'user_id':      s['id'],
                    'username':     s['username'],
                    'department':   s['department'],
                    'shift':        s['shift'],
                    'status':       status,
                    'current_page': act['current_page'],
                    'ip_addr':      act['ip_addr'],
                    'last_seen':    ago,
                    'last_active':  last_active,
                })
            else:
                result.append({
                    'user_id':      s['id'],
                    'username':     s['username'],
                    'department':   s['department'],
                    'shift':        s['shift'],
                    'status':       'offline',
                    'current_page': '—',
                    'ip_addr':      '—',
                    'last_seen':    'Never',
                    'last_active':  '—',
                })

        # Summary counts
        summary = {
            'active':  sum(1 for r in result if r['status'] == 'active'),
            'idle':    sum(1 for r in result if r['status'] == 'idle'),
            'offline': sum(1 for r in result if r['status'] == 'offline'),
            'total':   len(result),
        }

        return jsonify({'staff': result, 'summary': summary, 'updated': now.strftime('%H:%M:%S')})

    except Exception as e:
        logger.error(f"Activity fetch error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/activity')
def admin_activity():
    """Live activity monitor page"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))
    return render_template('activity.html')



# ---------------------------------------------------------------------------
# LEAVE REQUEST SYSTEM
# ---------------------------------------------------------------------------

LEAVE_TYPES = ['Casual Leave', 'Sick Leave', 'Earned Leave', 'Unpaid Leave']

@app.route('/staff/leave')
def staff_leave():
    """Staff leave request page"""
    if 'user_id' not in session or session['role'] != 'Staff':
        return redirect(url_for('login'))

    user_id = session['user_id']
    conn    = get_db_connection()

    my_requests = conn.execute(
        "SELECT * FROM leave_requests WHERE user_id=? ORDER BY applied_at DESC",
        (user_id,)
    ).fetchall()

    balance = get_leave_balance(user_id)
    conn.close()

    return render_template(
        'leave_staff.html',
        my_requests=my_requests,
        balance=balance,
        leave_types=LEAVE_TYPES,
    )


@app.route('/staff/leave/apply', methods=['POST'])
def apply_leave():
    """Staff submits a leave request"""
    if 'user_id' not in session or session['role'] != 'Staff':
        return redirect(url_for('login'))

    user_id    = session['user_id']
    username   = session['username']
    leave_type = request.form.get('leave_type', '').strip()
    from_date  = request.form.get('from_date', '').strip()
    to_date    = request.form.get('to_date', '').strip()
    reason     = request.form.get('reason', '').strip()

    if not all([leave_type, from_date, to_date, reason]):
        flash("All fields are required.")
        return redirect(url_for('staff_leave'))

    if leave_type not in LEAVE_TYPES:
        flash("Invalid leave type.")
        return redirect(url_for('staff_leave'))

    # Calculate days
    try:
        from datetime import datetime as dt
        fd = dt.strptime(from_date, '%Y-%m-%d')
        td = dt.strptime(to_date, '%Y-%m-%d')
        if td < fd:
            flash("End date cannot be before start date.")
            return redirect(url_for('staff_leave'))
        days = (td - fd).days + 1
    except ValueError:
        flash("Invalid date format.")
        return redirect(url_for('staff_leave'))

    conn = get_db_connection()

    # Check for overlapping pending/approved requests
    overlap = conn.execute(
        """SELECT id FROM leave_requests
           WHERE user_id=? AND status IN ('Pending','Approved')
           AND NOT (to_date < ? OR from_date > ?)""",
        (user_id, from_date, to_date)
    ).fetchone()

    if overlap:
        conn.close()
        flash("You already have a leave request for overlapping dates.")
        return redirect(url_for('staff_leave'))

    conn.execute(
        """INSERT INTO leave_requests
           (user_id, username, leave_type, from_date, to_date, days, reason, status, applied_at)
           VALUES (?,?,?,?,?,?,?,'Pending',?)""",
        (user_id, username, leave_type, from_date, to_date, days,
         reason, ist_now().strftime('%Y-%m-%d %H:%M:%S'))
    )

    # Notify all admins
    admins = conn.execute("SELECT id FROM users WHERE role='Admin'").fetchall()
    for admin in admins:
        push_notification(
            admin['id'],
            f"Leave Request — {username}",
            f"{username} applied for {leave_type} ({from_date} to {to_date}, {days} day{'s' if days>1 else ''}). Reason: {reason}",
            'warning',
            conn=conn
        )

    log_audit('LEAVE_APPLIED', f"{username}: {leave_type} {from_date}→{to_date}", user_id, conn=conn)
    conn.commit()
    conn.close()

    flash(f"✅ Leave request submitted for {days} day{'s' if days>1 else ''}. Awaiting approval.")
    return redirect(url_for('staff_leave'))


@app.route('/staff/leave/cancel/<int:req_id>', methods=['POST'])
def cancel_leave(req_id):
    """Staff cancels a pending leave request"""
    if 'user_id' not in session or session['role'] != 'Staff':
        return redirect(url_for('login'))

    user_id = session['user_id']
    conn    = get_db_connection()

    req = conn.execute(
        "SELECT * FROM leave_requests WHERE id=? AND user_id=?",
        (req_id, user_id)
    ).fetchone()

    if not req:
        flash("Request not found.")
        conn.close()
        return redirect(url_for('staff_leave'))

    if req['status'] != 'Pending':
        flash("Only pending requests can be cancelled.")
        conn.close()
        return redirect(url_for('staff_leave'))

    conn.execute("UPDATE leave_requests SET status='Cancelled' WHERE id=?", (req_id,))
    log_audit('LEAVE_CANCELLED', f"Request #{req_id}", user_id, conn=conn)
    conn.commit()
    conn.close()

    flash("✅ Leave request cancelled.")
    return redirect(url_for('staff_leave'))


@app.route('/admin/leave')
def admin_leave():
    """Admin leave management page"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    conn = get_db_connection()

    pending  = conn.execute(
        "SELECT * FROM leave_requests WHERE status='Pending' ORDER BY applied_at DESC"
    ).fetchall()
    history  = conn.execute(
        "SELECT * FROM leave_requests WHERE status!='Pending' ORDER BY applied_at DESC LIMIT 100"
    ).fetchall()

    # Leave balance for all staff
    staff = conn.execute(
        "SELECT id, username, full_name FROM users WHERE role='Staff' ORDER BY username"
    ).fetchall()

    balances = {}
    for s in staff:
        ensure_leave_balance(s['id'], conn)
        bal = conn.execute(
            "SELECT casual, sick, earned FROM leave_balance WHERE user_id=?",
            (s['id'],)
        ).fetchone()
        balances[s['id']] = dict(bal) if bal else {'casual':12,'sick':12,'earned':15}

    conn.commit()
    conn.close()

    return render_template(
        'leave_admin.html',
        pending=pending,
        history=history,
        staff=staff,
        balances=balances,
    )


@app.route('/admin/leave/action', methods=['POST'])
def admin_leave_action():
    """Admin approves or rejects leave request"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    req_id     = request.form.get('req_id', '')
    action     = request.form.get('action', '')       # 'approve' or 'reject'
    admin_note = request.form.get('admin_note', '').strip()

    if not req_id.isdigit() or action not in ('approve', 'reject'):
        flash("Invalid action.")
        return redirect(url_for('admin_leave'))

    conn = get_db_connection()
    req  = conn.execute(
        "SELECT * FROM leave_requests WHERE id=?", (int(req_id),)
    ).fetchone()

    if not req:
        flash("Leave request not found.")
        conn.close()
        return redirect(url_for('admin_leave'))

    if req['status'] != 'Pending':
        flash(f"This request is already {req['status']}.")
        conn.close()
        return redirect(url_for('admin_leave'))

    new_status  = 'Approved' if action == 'approve' else 'Rejected'
    reviewed_at = ist_now().strftime('%Y-%m-%d %H:%M:%S')
    reviewed_by = session['username']

    conn.execute(
        """UPDATE leave_requests
           SET status=?, admin_note=?, reviewed_at=?, reviewed_by=?
           WHERE id=?""",
        (new_status, admin_note, reviewed_at, reviewed_by, int(req_id))
    )

    if new_status == 'Approved':
        # Mark attendance as Leave for each date in the range
        from datetime import datetime as dt, timedelta
        fd = dt.strptime(req['from_date'], '%Y-%m-%d')
        td = dt.strptime(req['to_date'],   '%Y-%m-%d')
        d  = fd
        while d <= td:
            date_str = d.strftime('%Y-%m-%d')
            existing = conn.execute(
                "SELECT id FROM attendance WHERE user_id=? AND date=?",
                (req['user_id'], date_str)
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE attendance SET status='Leave', clock_in=NULL, clock_out=NULL WHERE id=?",
                    (existing['id'],)
                )
            else:
                conn.execute(
                    "INSERT INTO attendance (user_id, date, status) VALUES (?,?,?)",
                    (req['user_id'], date_str, 'Leave')
                )
            d += timedelta(days=1)

        # Deduct from leave balance
        balance_col = {
            'Casual Leave': 'casual',
            'Sick Leave':   'sick',
            'Earned Leave': 'earned',
        }.get(req['leave_type'])

        if balance_col:
            ensure_leave_balance(req['user_id'], conn)
            conn.execute(
                f"UPDATE leave_balance SET {balance_col}=MAX(0,{balance_col}-?) WHERE user_id=?",
                (req['days'], req['user_id'])
            )

        # Notify staff
        push_notification(
            req['user_id'],
            "Leave Approved ✅",
            f"Your {req['leave_type']} from {req['from_date']} to {req['to_date']} has been approved."
            + (f" Note: {admin_note}" if admin_note else ""),
            'success', conn=conn
        )
        log_audit('LEAVE_APPROVED',
                  f"#{req_id} {req['username']} {req['from_date']}→{req['to_date']}",
                  session['user_id'], conn=conn)
        flash(f"✅ Leave approved for {req['username']}. Attendance auto-marked.")

    else:
        # Notify staff of rejection
        push_notification(
            req['user_id'],
            "Leave Rejected ❌",
            f"Your {req['leave_type']} from {req['from_date']} to {req['to_date']} was rejected."
            + (f" Reason: {admin_note}" if admin_note else ""),
            'error', conn=conn
        )
        log_audit('LEAVE_REJECTED',
                  f"#{req_id} {req['username']} {req['from_date']}→{req['to_date']}",
                  session['user_id'], conn=conn)
        flash(f"❌ Leave rejected for {req['username']}.")

    conn.commit()
    conn.close()
    return redirect(url_for('admin_leave'))


@app.route('/admin/leave/balance', methods=['POST'])
def admin_leave_balance():
    """Admin manually adjusts leave balance"""
    if 'user_id' not in session or session['role'] != 'Admin':
        return redirect(url_for('login'))

    target_user_id = request.form.get('target_user_id', '')
    casual  = request.form.get('casual',  '12')
    sick    = request.form.get('sick',    '12')
    earned  = request.form.get('earned',  '15')

    if not target_user_id.isdigit():
        flash("Invalid user.")
        return redirect(url_for('admin_leave'))

    conn = get_db_connection()
    ensure_leave_balance(int(target_user_id), conn)
    conn.execute(
        "UPDATE leave_balance SET casual=?, sick=?, earned=? WHERE user_id=?",
        (int(casual), int(sick), int(earned), int(target_user_id))
    )
    log_audit('LEAVE_BALANCE_UPDATED',
              f"User #{target_user_id}: CL={casual} SL={sick} EL={earned}",
              session['user_id'], conn=conn)
    conn.commit()
    conn.close()
    flash("✅ Leave balance updated.")
    return redirect(url_for('admin_leave'))


# ---------------------------------------------------------------------------
# NOTIFICATION API
# ---------------------------------------------------------------------------

@app.route('/api/notifications')
def api_notifications():
    """Get notifications for current user"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    user_id = session['user_id']
    conn    = get_db_connection()

    notifs = conn.execute(
        """SELECT id, title, message, type, is_read, created_at
           FROM notifications WHERE user_id=?
           ORDER BY created_at DESC LIMIT 20""",
        (user_id,)
    ).fetchall()

    unread = conn.execute(
        "SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0",
        (user_id,)
    ).fetchone()[0]

    conn.close()

    return jsonify({
        'notifications': [dict(n) for n in notifs],
        'unread': unread,
    })


@app.route('/api/notifications/read', methods=['POST'])
def mark_notifications_read():
    """Mark all notifications as read"""
    if 'user_id' not in session:
        return jsonify({'ok': False}), 401

    conn = get_db_connection()
    conn.execute(
        "UPDATE notifications SET is_read=1 WHERE user_id=?",
        (session['user_id'],)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/notifications/read/<int:notif_id>', methods=['POST'])
def mark_one_read(notif_id):
    """Mark single notification as read"""
    if 'user_id' not in session:
        return jsonify({'ok': False}), 401

    conn = get_db_connection()
    conn.execute(
        "UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?",
        (notif_id, session['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/status')
def api_status():
    today       = today_str()
    server_time = ist_now().strftime('%H:%M:%S')
    conn = get_db_connection()
    total_staff = conn.execute("SELECT COUNT(*) FROM users WHERE role != 'Admin'").fetchone()[0]
    in_office   = conn.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=? AND clock_in IS NOT NULL AND clock_out IS NULL",
        (today,)
    ).fetchone()[0]
    conn.close()
    return jsonify({
        "system_status": "Healthy",
        "current_time_ist": server_time,
        "metrics": {
            "total_registered_staff": total_staff,
            "active_in_office_now": in_office,
        }
    })


# ---------------------------------------------------------------------------
# Auto-backup on startup
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# AUTO-AGENT LAUNCH (runs after staff login — no installation needed)
# ---------------------------------------------------------------------------

@app.route('/agent-launch')
def agent_launch():
    """Page shown after staff login — auto-triggers agent download"""
    if 'user_id' not in session or session['role'] != 'Staff':
        return redirect(url_for('login'))
    return render_template('agent_launch.html')


@app.route('/download-agent')
def download_agent():
    """Serve a personalized .bat file — fixes VBS pythonw path issue"""
    if 'user_id' not in session:
        return redirect(url_for('login'))

    username  = session['username']
    server    = f"{request.scheme}://{request.host}"
    url       = f"{server}/agent-script?u={username}"

    # Build bat using triple-quoted string — avoids escaping issues
    bat_lines = [
        "@echo off",
        f":: StaffPortal Agent for {username}",
        "setlocal EnableDelayedExpansion",
        "set AGENT_DIR=%APPDATA%\\StaffAgent",
        "mkdir \"%AGENT_DIR%\" 2>nul",
        "",
        ":: Find Python full path",
        "set PYTHONW=",
        "for /f \"delims=\" %%i in (\'where pythonw.exe 2^>nul\') do if \"!PYTHONW!\"==\"\" set PYTHONW=%%i",
        "if \"%PYTHONW%\"==\"\" for /f \"delims=\" %%i in (\'where python.exe 2^>nul\') do if \"!PYTHONW!\"==\"\" set PYTHONW=%%i",
        "if \"%PYTHONW%\"==\"\" (",
        "  echo Python not found. Install Python 3 first.",
        "  pause",
        "  exit /b 1",
        ")",
        "",
        ":: Download agent script from server",
        f'powershell -Command \"(New-Object Net.WebClient).DownloadFile(\'{url}\', \'%AGENT_DIR%\\\\agent.py\')\"',
        "if not exist \"%AGENT_DIR%\\agent.py\" (",
        "  echo Download failed. Check network.",
        "  pause",
        "  exit /b 1",
        ")",
        "",
        ":: Write VBS launcher using full Python path",
        "(",
        "echo Set objShell = WScript.CreateObject^(\"WScript.Shell\"^)",
        "echo objShell.Run Chr^(34^) ^& \"%PYTHONW%\" ^& Chr^(34^) ^& \" \" ^& Chr^(34^) ^& \"%AGENT_DIR%\\agent.py\" ^& Chr^(34^), 0, False",
        ") > \"%AGENT_DIR%\\launch.vbs\"",
        "",
        ":: Add to Windows Startup",
        "copy /y \"%AGENT_DIR%\\launch.vbs\" \"%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\Startup\\StaffPortalAgent.vbs\" >nul",
        "",
        ":: Start agent now",
        "wscript.exe \"%AGENT_DIR%\\launch.vbs\"",
        "exit",
    ]
    bat = "\r\n".join(bat_lines)

    from flask import Response
    return Response(
        bat,
        mimetype='application/octet-stream',
        headers={'Content-Disposition': f'attachment; filename=start_agent_{username}.bat'}
    )


@app.route('/agent-script')
def agent_script():
    """Serve personalized agent.py with credentials embedded"""
    if 'user_id' not in session:
        # Allow download via URL param as fallback
        u = request.args.get('u', '')
        if not u:
            return "Unauthorized", 401
    else:
        u = session.get('username', request.args.get('u', ''))

    server_ip = f"{request.scheme}://{request.host}"

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username=?", (u,)).fetchone()
    conn.close()

    if not user:
        return "User not found", 404

    # Read base agent template and embed credentials
    agent_template_path = os.path.join(os.path.dirname(__file__), 'agent', 'agent.py')
    if not os.path.exists(agent_template_path):
        return "Agent template not found", 404

    with open(agent_template_path, 'r') as f:
        content = f.read()

    # Embed server URL and username
    content = content.replace('http://192.168.1.100:5000', server_ip)
    content = content.replace('"john"', f'"{u}"')

    from flask import Response
    return Response(content, mimetype='text/plain')


# ---------------------------------------------------------------------------
# AUTO LATE / ABSENT SCHEDULER
# ---------------------------------------------------------------------------

import threading

def parse_shift_start(shift_str):
    """Parse shift string like '09:00 AM - 06:00 PM' → datetime.time"""
    try:
        if not shift_str or shift_str.lower() in ('flexible', 'night shift'):
            return None
        start_part = shift_str.split('-')[0].strip()  # e.g. "09:00 AM"
        from datetime import datetime as dt
        return dt.strptime(start_part, '%I:%M %p').time()
    except Exception:
        return None


def run_attendance_checker():
    """
    Background thread: runs every 60 seconds.
    - Clears stale Late/Absent if weekoff was changed to today (Bug 1 fix)
    - Only logs/updates once per status transition (Bug 2 fix)
    - Guarantees conn.close() via finally (Bug 4 fix)
    - Respects Leave — never auto-overrides admin manual marks
    """
    import time as time_module

    while True:
        conn = None
        try:
            now_ist   = ist_now()
            today     = now_ist.strftime('%Y-%m-%d')
            today_day = now_ist.strftime('%A')
            now_time  = now_ist.time()

            conn  = get_db_connection()
            staff = conn.execute(
                "SELECT id, username, shift, weekoff FROM users WHERE role='Staff'"
            ).fetchall()

            for s in staff:
                # If today is their weekoff — clear any stale auto Late/Absent
                # (handles: weekoff changed AFTER scheduler already marked them)
                if s['weekoff'] == today_day:
                    stale = conn.execute(
                        "SELECT id, status FROM attendance WHERE user_id=? AND date=?",
                        (s['id'], today)
                    ).fetchone()
                    if stale and stale['status'] in ('Late', 'Absent'):
                        conn.execute("DELETE FROM attendance WHERE id=?", (stale['id'],))
                        log_audit('AUTO_WEEKOFF_CLEARED',
                                  f"{s['username']} — cleared stale {stale['status']}",
                                  s['id'], conn=conn)
                        logger.info(f"WEEKOFF-RECONCILE: cleared {stale['status']} for {s['username']}")
                    continue

                shift_start = parse_shift_start(s['shift'])
                if not shift_start:
                    continue

                from datetime import datetime as dt
                shift_dt  = dt.combine(now_ist.date(), shift_start)
                now_dt    = dt.combine(now_ist.date(), now_time)
                mins_late = (now_dt - shift_dt).total_seconds() / 60

                if mins_late < 0:
                    continue  # Shift hasn't started yet

                record = conn.execute(
                    "SELECT * FROM attendance WHERE user_id=? AND date=?",
                    (s['id'], today)
                ).fetchone()

                if record and record['clock_in']:
                    continue  # Already clocked in

                if record and record['status'] == 'Leave':
                    continue  # Admin-marked Leave — never auto-override

                if mins_late >= 60:
                    # Only act if not already Absent (avoids spam)
                    if not record or record['status'] != 'Absent':
                        if record:
                            conn.execute(
                                "UPDATE attendance SET status='Absent' WHERE user_id=? AND date=?",
                                (s['id'], today)
                            )
                        else:
                            conn.execute(
                                "INSERT INTO attendance (user_id, date, status) VALUES (?,?,?)",
                                (s['id'], today, 'Absent')
                            )
                        log_audit('AUTO_ABSENT',
                                  f"{s['username']} — {mins_late:.0f} mins past shift start",
                                  s['id'], conn=conn)
                        logger.info(f"AUTO-ABSENT: {s['username']} ({mins_late:.0f} mins late)")

                elif mins_late >= 30:
                    # Only act if not already Late or Absent (avoids spam)
                    if not record or record['status'] not in ('Late', 'Absent'):
                        if record:
                            conn.execute(
                                "UPDATE attendance SET status='Late' WHERE user_id=? AND date=?",
                                (s['id'], today)
                            )
                        else:
                            conn.execute(
                                "INSERT INTO attendance (user_id, date, status) VALUES (?,?,?)",
                                (s['id'], today, 'Late')
                            )
                        log_audit('AUTO_LATE',
                                  f"{s['username']} — {mins_late:.0f} mins past shift start",
                                  s['id'], conn=conn)
                        logger.info(f"AUTO-LATE: {s['username']} ({mins_late:.0f} mins late)")

            conn.commit()

        except Exception as e:
            logger.error(f"Attendance checker error: {e}")
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

        time_module.sleep(60)


def start_autobackup():
    """Create backup on startup"""
    import atexit
    import signal
    
    def backup_handler(signum=None, frame=None):
        logger.info("Creating backup before shutdown...")
        create_backup()
        cleanup_old_backups()
    
    atexit.register(backup_handler)
    signal.signal(signal.SIGTERM, backup_handler)
    signal.signal(signal.SIGINT, backup_handler)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    init_db()
    start_autobackup()

    # Guard against Werkzeug reloader double-spawning the scheduler thread.
    # WERKZEUG_RUN_MAIN is only 'true' in the actual worker child process.
    debug_mode       = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    is_reloader_main = os.environ.get('WERKZEUG_RUN_MAIN') == 'true'

    if not debug_mode or is_reloader_main:
        checker = threading.Thread(target=run_attendance_checker, daemon=True)
        checker.start()
        logger.info("✅ Auto Late/Absent/Weekoff scheduler started")
    else:
        logger.info("Skipping scheduler in reloader parent process")

    app.run(host='0.0.0.0', port=5000, debug=debug_mode)
